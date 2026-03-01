#!/usr/bin/env python3
"""
财务原始数据导入工具 - Python版本
独立运行，无需PHP环境，直接操作数据库
"""

import sys
import os
from pathlib import Path
import pymysql
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import re
from typing import Dict, List, Optional, Tuple
import argparse


class PaymentImporter:
    """财务数据导入器"""
    
    # 颜色定义（RGB格式）
    COLORS = {
        'blue': 'FF99CCFF',      # 蓝色
        'pink': 'FFFFCCFF',      # 粉红色  
        'green': 'FFCCFFCC',     # 绿色
    }
    
    def __init__(self, db_config: Dict):
        """初始化导入器"""
        self.db_config = db_config
        self.conn = None
        
    def connect_db(self):
        """连接数据库"""
        self.conn = pymysql.connect(
            host=self.db_config['host'],
            port=self.db_config.get('port', 3306),
            user=self.db_config['user'],
            password=self.db_config['password'],
            database=self.db_config['database'],
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        
    def close_db(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            
    def parse_excel(self, file_path: str) -> List[Dict]:
        """解析Excel文件"""
        print(f"解析Excel文件: {file_path}")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        rows = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"  处理工作表: {sheet_name}")
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # 获取单元格颜色
                fill = row[0].fill
                color = self._get_color_name(fill)
                
                # 提取数据
                member_id = self._extract_member_id(row[0].value)
                if not member_id:
                    continue
                    
                row_data = {
                    'row_number': row_idx,
                    'sheet_name': sheet_name,
                    'member_id': member_id,
                    'color': color,
                    'raw_value': str(row[0].value) if row[0].value else ''
                }
                
                rows.append(row_data)
        
        print(f"解析完成，共 {len(rows)} 行数据\n")
        return rows
        
    def _get_color_name(self, fill: PatternFill) -> str:
        """获取颜色名称"""
        if fill.patternType == 'solid' and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb:
                for name, color_code in self.COLORS.items():
                    if rgb == color_code:
                        return name
        return 'unknown'
        
    def _extract_member_id(self, value) -> Optional[str]:
        """提取Member_ID"""
        if not value:
            return None
            
        value_str = str(value).strip()
        
        # 匹配各种格式
        patterns = [
            r'Member_ID[:\s]*(\d+)',
            r'(\d{6,})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, value_str, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return None
        
    def match_orders(self, rows: List[Dict], wxapp_id: int) -> List[Dict]:
        """匹配订单"""
        print("匹配订单...")
        
        # 提取所有member_id
        member_ids = [row['member_id'] for row in rows]
        
        if not member_ids:
            return rows
            
        # 批量查询订单
        cursor = self.conn.cursor()
        placeholders = ','.join(['%s'] * len(member_ids))
        
        sql = f"""
            SELECT id, order_sn, member_id, is_pay, pay_time
            FROM yoshop_inpack
            WHERE wxapp_id = %s AND member_id IN ({placeholders})
        """
        
        cursor.execute(sql, [wxapp_id] + member_ids)
        orders = cursor.fetchall()
        cursor.close()
        
        # 按member_id分组
        orders_by_member = {}
        for order in orders:
            mid = str(order['member_id'])
            if mid not in orders_by_member:
                orders_by_member[mid] = []
            orders_by_member[mid].append(order)
        
        # 匹配订单
        for row in rows:
            mid = row['member_id']
            matched = orders_by_member.get(mid, [])
            
            row['matched_orders'] = matched
            row['match_count'] = len(matched)
            
            if len(matched) == 1:
                row['matched_order'] = matched[0]
                row['match_status'] = 'single'
            elif len(matched) > 1:
                row['match_status'] = 'multiple'
            else:
                row['match_status'] = 'none'
        
        # 统计
        stats = {
            'total': len(rows),
            'matched': sum(1 for r in rows if r['match_count'] == 1),
            'unmatched': sum(1 for r in rows if r['match_count'] == 0),
            'multiple': sum(1 for r in rows if r['match_count'] > 1),
            'blue': sum(1 for r in rows if r['color'] == 'blue'),
            'pink': sum(1 for r in rows if r['color'] == 'pink'),
            'green': sum(1 for r in rows if r['color'] == 'green'),
            'unknown': sum(1 for r in rows if r['color'] == 'unknown'),
        }
        
        print(f"匹配完成:")
        print(f"  总行数: {stats['total']}")
        print(f"  匹配成功: {stats['matched']}")
        print(f"  未匹配: {stats['unmatched']}")
        print(f"  多重匹配: {stats['multiple']}")
        print(f"  蓝色: {stats['blue']}, 粉红: {stats['pink']}, 绿色: {stats['green']}, 未知: {stats['unknown']}\n")
        
        return rows
        
    def execute_import(self, rows: List[Dict], dry_run: bool = False) -> Dict:
        """执行导入"""
        print(f"执行导入 (dry_run={dry_run})...")
        
        success_count = 0
        failure_count = 0
        failed_items = []
        
        cursor = self.conn.cursor()
        
        for row in rows:
            # 只处理蓝色和粉红色的单一匹配行
            if row['color'] not in ['blue', 'pink']:
                continue
                
            if row['match_status'] != 'single':
                continue
            
            order = row['matched_order']
            order_id = order['id']
            
            try:
                if not dry_run:
                    # 更新订单支付状态
                    sql = """
                        UPDATE yoshop_inpack
                        SET is_pay = 1, pay_time = %s
                        WHERE id = %s
                    """
                    cursor.execute(sql, [int(datetime.now().timestamp()), order_id])
                
                success_count += 1
                print(f"  ✓ 订单 #{order_id} ({order['order_sn']}) - Member_ID: {row['member_id']}")
                
            except Exception as e:
                failure_count += 1
                failed_items.append({
                    'member_id': row['member_id'],
                    'order_id': order_id,
                    'error': str(e)
                })
                print(f"  ✗ 订单 #{order_id} 失败: {e}")
        
        if not dry_run:
            self.conn.commit()
        
        cursor.close()
        
        result = {
            'success': failure_count == 0,
            'total_processed': success_count + failure_count,
            'success_count': success_count,
            'failure_count': failure_count,
            'failed_items': failed_items
        }
        
        print(f"\n导入完成:")
        print(f"  成功: {success_count}")
        print(f"  失败: {failure_count}")
        
        return result


def main():
    parser = argparse.ArgumentParser(description='财务原始数据导入工具')
    parser.add_argument('excel_file', help='Excel文件路径')
    parser.add_argument('--wxapp-id', type=int, default=10022, help='小程序ID')
    parser.add_argument('--dry-run', action='store_true', help='试运行模式（不实际更新数据库）')
    parser.add_argument('--db-host', default='localhost', help='数据库主机')
    parser.add_argument('--db-port', type=int, default=3306, help='数据库端口')
    parser.add_argument('--db-user', required=True, help='数据库用户名')
    parser.add_argument('--db-password', required=True, help='数据库密码')
    parser.add_argument('--db-name', required=True, help='数据库名称')
    
    args = parser.parse_args()
    
    # 检查文件
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件不存在: {args.excel_file}")
        sys.exit(1)
    
    # 数据库配置
    db_config = {
        'host': args.db_host,
        'port': args.db_port,
        'user': args.db_user,
        'password': args.db_password,
        'database': args.db_name,
    }
    
    print("=== 财务原始数据导入工具 ===\n")
    print(f"Excel文件: {args.excel_file}")
    print(f"小程序ID: {args.wxapp_id}")
    print(f"模式: {'试运行' if args.dry_run else '正式导入'}\n")
    
    # 执行导入
    importer = PaymentImporter(db_config)
    
    try:
        importer.connect_db()
        
        # 解析Excel
        rows = importer.parse_excel(args.excel_file)
        
        # 匹配订单
        rows = importer.match_orders(rows, args.wxapp_id)
        
        # 执行导入
        result = importer.execute_import(rows, dry_run=args.dry_run)
        
        if result['success']:
            print("\n✓ 导入成功!")
            sys.exit(0)
        else:
            print("\n✗ 导入部分失败")
            sys.exit(1)
            
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
        
    finally:
        importer.close_db()


if __name__ == '__main__':
    main()
